#!python
# -*- coding: utf-8 -*-

import sys
import time
import glob
import csv
import codecs
import xlrd
import xlwt
import openpyxl
#import pandas as pd
import xlsxwriter
from openpyxl.styles import Font, Color, colors, PatternFill

READNAME = "Daily.xlsx"
SAVENAME = "consult.xlsx"
#配列関連
#二次元でもよかったか、indexと中身が混乱しやすいと思い、独立の配列にした。
listKey = []
listASIN = []
listSeller = []
listName = []
listType = []
listStock = []
dailyList = []
folderName = "today"
renewDate = ""
readsize=0
indexAdd=0

def csvHelper():
    #====================csv
    print "please input the [floder] name"
    #folderName = raw_input()
    global folderName
    global renewDate
    print "the whole file in folder \"" + folderName + " \" will be combined as a file"

    files = glob.glob('./'+folderName+'/*')
    if(len(files)==0):
        print "file don's exist!!"
    
    for file in files:
        print "reading csv file:" + file
        renewDate = files[0][8:16]
        f = open(file,'r')

        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            i = 0
            key = ""
            Asin = ""
            Seller = ""
            Name = ""
            Type = ""
    
            for item in row:
                if i==4:
                    key+=str(item)
                    Asin += str(item)
                elif i == 6:
                    key+=str(item)
                    Seller+=str(item)
                elif i == 7:
                    Name+=str(item)
                elif i == 12:
                    Type+=str(item)
                elif i == 17:
                    listStock.append(int(item))
                i=i+1
            listASIN.append(Asin)
            listSeller.append(Seller)
            listName.append(Name)
            listType.append(Type)
            listKey.append(key)
        f.close()

def createTodayReport():
    workbook = xlsxwriter.Workbook(renewDate+'.xlsx')
    worksheet = workbook.add_worksheet()


    for row in range(len(listASIN)):
        for column in range(6):
            if(column==0):
                worksheet.write(row,column,listKey[row].decode('shift-jis'))
            if(column==1):
                worksheet.write(row,column,listASIN[row].decode('shift-jis'))
            if(column==2):
                worksheet.write(row,column,listSeller[row].decode('shift-jis'))
            if(column==3):
                worksheet.write(row,column,listName[row].decode('shift-jis'))
            if(column==4):
                worksheet.write(row,column,listType[row].decode('shift-jis'))
            if(column==5):
                worksheet.write(row,column,listStock[row])
    workbook.close()

def makeDailyReport():
    global readsize
    #書き込み不可が、中身を文字列として見れる
    readbook = xlrd.open_workbook(READNAME)
    readsheet = readbook.sheet_by_index(0)
    global indexAdd
    indexAdd = readsheet.ncols  #0-indexAdd

    #セルに対するマクロ操作が可能
    workbook = openpyxl.load_workbook(READNAME)
    worksheet = workbook.get_sheet_by_name('Sheet1')
    
    #全体を探索し、同じキーのやつを記載更新する。
    for i in range(readsheet.nrows):
        print "writing data" + str(i)
        key = readsheet.cell(rowx=i, colx=0).value.encode('cp932')
        indexHit = 0
        try:
            indexHit = listKey.index(key)
        except ValueError:
            continue
        lastStock = 0
        soldStock = 0

        if(worksheet.cell(row=i+1,column=indexAdd-1).value != None):
            lastStock = worksheet.cell(row=i+1, column = indexAdd-1).value
            soldStock = lastStock-listStock[indexHit]
        worksheet.cell(row=i+1,column = indexAdd+1).value = listStock[indexHit]
        worksheet.cell(row=i+1,column = indexAdd+2).value = soldStock
        if(soldStock<=0):#green
            worksheet.cell(row=i+1,column = indexAdd+2).fill = PatternFill(patternType='solid', fgColor='FFC0FFC0')
        else:#pink
            worksheet.cell(row=i+1,column = indexAdd+2).fill = PatternFill(patternType='solid', fgColor='FFFFC0C0')

        del listKey[indexHit]
        del listASIN[indexHit]
        del listSeller[indexHit]
        del listName[indexHit]
        del listType[indexHit]
        del listStock[indexHit]

    
    for i in range(len(listASIN)):
        if(len(listASIN[i])<8):
            continue
        else:
            worksheet.cell(row=i+1+readsheet.nrows,column = 1).value = listKey[i].decode('shift-jis')
            worksheet.cell(row=i+1+readsheet.nrows,column = 2).value = listASIN[i].decode('shift-jis')
            worksheet.cell(row=i+1+readsheet.nrows,column = 3).value = listSeller[i].decode('shift-jis')
            try:
                worksheet.cell(row=i+1+readsheet.nrows,column = 4).value = listName[i].decode('shift-jis')
            except:
                worksheet.cell(row=i+1+readsheet.nrows,column = 4).value = "Name error!".decode('shift-jis')
            worksheet.cell(row=i+1+readsheet.nrows,column = 5).value = listType[i].decode('shift-jis')
            worksheet.cell(row=i+1+readsheet.nrows,column = indexAdd+1).value =listStock[i]

    print u"更新する日付は:"+renewDate
    print
    print u"今日のデータを日報に書き込んでいます"
    print u"しばらくこのままお待ちください。。。"
    
    worksheet.cell(row = 1,column = indexAdd+1).value = renewDate.decode('utf-8')
    readsize=readsheet.nrows
    workbook.save(SAVENAME)




    
if __name__ == '__main__':
    print u"アプリ起動中"

    csvHelper()
    print
    print u"今日のデータファイルを作成しています。"
    print u"このまましばらくお待ちください。。。"
    print
    createTodayReport()
    makeDailyReport()
    print len(listASIN)
    time.sleep(2.0)
    
    print u"success!!すべて完了しました。エンターキーを押して終了してください。"

    raw_input()
